import os
from eset_status import ESETStatusChecker
from wsus_status import WSUSStatusChecker
from ocs_status import OCSStatusChecker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

# Lista de equipos a ignorar (lista negra)
blacklist = [
    'server1', 'server2', 'monitor', 'backup', 'virtual1', 'virtual2',
    'cloud', 'sync', 'fileserver', 'inventory', 'backupserver', 'support1',
    'support2', 'main', 'domaincontroller', 'updateserver'
]

def normalize_name(name):
    if name.endswith(".midominio.local"):
        return name[:-15].lower()  # Eliminar ".midominio.local" y convertir a minúsculas
    return name.lower()

def merge_status_tables(eset_table, wsus_table, ocs_table):
    eset_dict = {normalize_name(row['Nombre']): row['Estado'] for row in eset_table}
    wsus_dict = {normalize_name(row['Nombre']): row['Estado'] for row in wsus_table}
    ocs_computers = {normalize_name(row['Computer']): row['Computer'] for row in ocs_table}

    merged_data = []

    all_names = set(eset_dict.keys()).union(set(wsus_dict.keys()))

    for name in all_names:
        if name not in blacklist:  # Ignorar los nombres en la lista negra
            ocs_status = "Mal"
            for ocs_name in ocs_computers.values():
                if name == normalize_name(ocs_name) or name.startswith(normalize_name(ocs_name)[:15]):
                    ocs_status = "OK"
                    break
            merged_data.append({
                'Nombre': name,
                'ESET': eset_dict.get(name, 'N/A'),
                'WSUS': wsus_dict.get(name, 'N/A'),
                'OCS': ocs_status
            })

    return merged_data

def create_xlsx(data, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Status Report"

    # Definir colores
    fill_ok = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_mal = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_na = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    # Definir bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header = ['Nombre', 'ESET', 'WSUS', 'OCS']
    sheet.append(header)

    # Aplicar bordes a la cabecera
    for col_num, _ in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.border = thin_border

    for row in data:
        row_data = [row['Nombre'], row['ESET'], row['WSUS'], row['OCS']]
        sheet.append(row_data)
        for col_num, value in enumerate(row_data[1:], start=2):  # Empezar desde la columna 2 para 'ESET', 'WSUS', y 'OCS'
            cell = sheet.cell(row=sheet.max_row, column=col_num)
            cell.border = thin_border
            if value == 'OK':
                cell.fill = fill_ok
            elif value == 'Mal':
                cell.fill = fill_mal
            elif value == 'N/A':
                cell.fill = fill_na

    # Aplicar bordes a todas las celdas
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    workbook.save(output_path)

def main():
    eset_checker = ESETStatusChecker()
    wsus_checker = WSUSStatusChecker()
    ocs_checker = OCSStatusChecker()

    eset_status_table = eset_checker.run()
    wsus_status_table = wsus_checker.get_status_table()
    ocs_status_table = ocs_checker.run()

    merged_data = merge_status_tables(eset_status_table, wsus_status_table, ocs_status_table)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'status_report.xlsx')
    create_xlsx(merged_data, output_path)

    print(f"Status report generated and saved to {output_path}")

if __name__ == "__main__":
    main()